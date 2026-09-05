import io
import csv
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from app.database import create_post
from app.time_utils import normalize_schedule

COLUMNS = [
    ("title", "Tiêu đề / Chiến dịch (title)", 25),
    ("fb_caption", "Nội dung Facebook (fb_caption)", 40),
    ("ig_caption", "Nội dung Instagram (ig_caption)", 35),
    ("google_caption", "Nội dung Google Maps (google_caption)", 35),
    ("google_action_type", "Nút bấm Google (LEARN_MORE / ORDER / CALL / BOOK / NONE)", 22),
    ("google_action_url", "Link nút Google (google_action_url)", 30),
    ("images", "Hình ảnh (images - cách nhau dấu phẩy)", 32),
    ("target_fb", "Đăng FB (1/0)", 14),
    ("target_ig", "Đăng IG (1/0)", 14),
    ("target_story", "Đăng Story 9:16 (1/0)", 18),
    ("target_google", "Đăng Google (1/0)", 16),
    ("scheduled_time", "Ngày giờ đăng (YYYY-MM-DD HH:mm)", 25)
]

def map_action_type(val: str) -> str:
    """Normalize Google Action Type from Vietnamese or English text."""
    if not val:
        return "LEARN_MORE"
    v = str(val).strip().upper()
    v_lower = str(val).strip().lower()
    
    if "ORDER" in v or "đặt hàng" in v_lower or "đặt món" in v_lower or "mua" in v_lower:
        return "ORDER"
    elif "CALL" in v or "gọi" in v_lower or "hotline" in v_lower:
        return "CALL"
    elif "BOOK" in v or "đặt lịch" in v_lower or "đặt bàn" in v_lower:
        return "BOOK"
    elif "SIGN_UP" in v or "đăng ký" in v_lower:
        return "SIGN_UP"
    elif "NONE" in v or "không" in v_lower or "tắt" in v_lower:
        return "NONE"
    else:
        return "LEARN_MORE"

def generate_bulk_excel_template() -> io.BytesIO:
    wb = openpyxl.Workbook()
    
    # ----------------------------------------------------
    # SHEET 1: DANH SÁCH ĐĂNG BÀI
    # ----------------------------------------------------
    ws1 = wb.active
    ws1.title = "Danh Sách Đăng Bài"

    # Header style
    header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid") # Navy blue
    header_font = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    
    thin_border = Border(
        left=Side(style='thin', color='D1D5DB'),
        right=Side(style='thin', color='D1D5DB'),
        top=Side(style='thin', color='D1D5DB'),
        bottom=Side(style='thin', color='D1D5DB')
    )

    # Write Headers
    for col_idx, (col_key, col_title, col_width) in enumerate(COLUMNS, 1):
        cell = ws1.cell(row=1, column=col_idx, value=col_title)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = align_center
        cell.border = thin_border
        ws1.column_dimensions[get_column_letter(col_idx)].width = col_width

    ws1.row_dimensions[1].height = 32

    # Sample rows
    now = datetime.now()
    sample_data = [
        (
            "Măng cụt Bảo Lộc Hữu Cơ",
            "🌿 MĂNG CỤT BẢO LỘC – NGỌT LÀNH TỪ VÙNG ĐẤT CAO NGUYÊN\n\nMùa măng cụt chín cây thơm ngon đã chính thức cập bến ROOTS! Từng trái vỏ mỏng, ruột trắng nõn, vị chua ngọt thanh mát chuẩn hữu cơ.\n\n👉 Ghé ROOTS hoặc đặt hàng ngay hôm nay!",
            "Măng cụt Bảo Lộc chín cây chuẩn hữu cơ cập bến ROOTS ✨ Vị ngọt thanh thanh, từng múi trắng ngần mọng nước. #roots #organic #mangcut #fruits #saigon",
            "Măng cụt Bảo Lộc canh tác hữu cơ đã có mặt tại cửa hàng ROOTS. Thưởng thức vị ngọt thanh mát tự nhiên ngay hôm nay!",
            "ORDER",
            "https://roots.vn",
            "mangcut1.jpg, mangcut2.jpg",
            1, 1, 1, 1,
            (now + timedelta(days=1)).strftime("%Y-%m-%d 09:00")
        ),
        (
            "Cold Pressed Juice Healthy",
            "🍹 NẠP NĂNG LƯỢNG TƯƠI MỚI VỚI NƯỚC ÉP NGUYÊN CHẤT ROOTS\n\n100% rau củ quả organic ép lạnh không thêm đường hay nước. Giúp thanh lọc cơ thể và tăng cường đề kháng cho ngày dài làm việc!",
            "100% Cold-Pressed Juice tươi mát mỗi ngày 🌿 Healthy & Fresh vibes at ROOTS! #juice #detox #healthy #organicfood",
            "Nước ép rau củ quả ép lạnh nguyên chất ROOTS - Thanh lọc cơ thể, trọn vẹn vitamin.",
            "LEARN_MORE",
            "https://roots.vn/menu-juice",
            "juice1.jpg",
            1, 1, 1, 1,
            (now + timedelta(days=2)).strftime("%Y-%m-%d 15:00")
        ),
        (
            "Khuyến mãi Cuối Tuần",
            "🎉 DEAL HOT CUỐI TUẦN: GIẢM 15% CHO TOÀN BỘ SẢN PHẨM HỮU CƠ\n\nÁp dụng cho mọi hóa đơn mua sắm trực tiếp tại cửa hàng hoặc đặt online qua website ROOTS.",
            "Weekend Deal Alert! 🔥 Giảm ngay 15% toàn bộ sản phẩm hữu cơ tại ROOTS. Ghé ngay nhé cả nhà! #sale #weekenddeal #organicstore",
            "Ưu đãi cuối tuần: Giảm 15% toàn bộ thực phẩm hữu cơ tại ROOTS - 237-239 Nguyễn Công Trứ, Q1.",
            "ORDER",
            "https://roots.vn",
            "promo.jpg",
            1, 1, 1, 1,
            (now + timedelta(days=3)).strftime("%Y-%m-%d 10:30")
        )
    ]

    for row_idx, row_values in enumerate(sample_data, 2):
        for col_idx, val in enumerate(row_values, 1):
            cell = ws1.cell(row=row_idx, column=col_idx, value=val)
            cell.font = Font(name="Segoe UI", size=9.5)
            cell.border = thin_border
            # Alignment
            if col_idx in [1, 2, 3, 4, 6, 7]:
                cell.alignment = align_left
            else:
                cell.alignment = align_center
        ws1.row_dimensions[row_idx].height = 28

    # ----------------------------------------------------
    # SHEET 2: HƯỚNG DẪN CHI TIẾT
    # ----------------------------------------------------
    ws2 = wb.create_sheet(title="Hướng Dẫn Sử Dụng")
    
    guide_title_font = Font(name="Segoe UI", size=14, bold=True, color="1E3A8A")
    ws2["A1"] = "HƯỚNG DẪN ĐIỀN DỮ LIỆU ĐĂNG BÀI TỰ ĐỘNG HÀNG LOẠT"
    ws2["A1"].font = guide_title_font
    
    guide_headers = ["Tên Cột", "Mô Tả & Quy Tắc", "Ví Dụ Mẫu"]
    for col_i, gh in enumerate(guide_headers, 1):
        c = ws2.cell(row=3, column=col_i, value=gh)
        c.fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
        c.font = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
        c.alignment = align_center
        c.border = thin_border
    
    guide_rows = [
        ("title", "Tên chiến dịch hoặc tiêu đề nội bộ để bạn dễ theo dõi trong bảng danh sách.", "Măng cụt Bảo Lộc"),
        ("fb_caption", "Nội dung bài viết Facebook. Có thể dùng icon, xuống dòng thoải mái.", "🌿 Măng cụt tươi ngon vừa về..."),
        ("ig_caption", "Nội dung Instagram (kèm #hashtag). Nếu để trống, hệ thống sẽ tự động lấy từ cột Facebook.", "Măng cụt chín cây #roots #fruits"),
        ("google_caption", "Nội dung đăng lên Google Maps & Search. Tối đa 1500 ký tự. Nếu để trống, tự lấy từ cột Facebook.", "Măng cụt Bảo Lộc tại ROOTS"),
        ("google_action_type", "Loại nút CTA trên Google: LEARN_MORE (Tìm hiểu thêm), ORDER (Đặt hàng), CALL (Gọi ngay), BOOK (Đặt bàn), NONE (Không nút).", "ORDER"),
        ("google_action_url", "Đường dẫn website khi khách bấm vào nút trên Google Maps (Không áp dụng nếu chọn CALL hoặc NONE).", "https://roots.vn"),
        ("images", "Tên file ảnh đã tải lên hoặc link ảnh web công khai (cách nhau bởi dấu phẩy nếu đăng nhiều ảnh).", "anh1.jpg, anh2.jpg"),
        ("target_fb", "1 = Đăng lên Facebook Fanpage, 0 = Không đăng.", "1"),
        ("target_ig", "1 = Đăng lên Instagram Business, 0 = Không đăng.", "1"),
        ("target_story", "1 = Tự động tạo Story 9:16 và đăng lên FB & IG Story, 0 = Không đăng Story.", "1"),
        ("target_google", "1 = Đăng lên Google Maps & Search (ROOTS), 0 = Không đăng Google.", "1"),
        ("scheduled_time", "Ngày giờ đăng bài. Định dạng: YYYY-MM-DD HH:mm (Ví dụ: 2026-08-25 09:30) hoặc DD/MM/YYYY HH:mm.", "2026-08-25 09:30")
    ]
    
    for r_idx, (gc1, gc2, gc3) in enumerate(guide_rows, 4):
        c1 = ws2.cell(row=r_idx, column=1, value=gc1)
        c2 = ws2.cell(row=r_idx, column=2, value=gc2)
        c3 = ws2.cell(row=r_idx, column=3, value=gc3)
        for c in [c1, c2, c3]:
            c.font = Font(name="Segoe UI", size=9.5)
            c.border = thin_border
            c.alignment = align_left
        c1.font = Font(name="Segoe UI", size=9.5, bold=True, color="1E40AF")
        ws2.row_dimensions[r_idx].height = 24

    ws2.column_dimensions["A"].width = 22
    ws2.column_dimensions["B"].width = 65
    ws2.column_dimensions["C"].width = 35

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output

def parse_bulk_file(file_content: bytes, filename: str) -> list:
    posts = []
    
    if filename.lower().endswith(".csv"):
        # Parse CSV
        text_content = file_content.decode("utf-8-sig", errors="ignore")
        reader = csv.reader(io.StringIO(text_content))
        rows = list(reader)
        if not rows:
            return []
        data_rows = rows[1:]
        for r in data_rows:
            if not r or not any(str(x).strip() for x in r):
                continue
            
            title = r[0].strip() if len(r) > 0 else ""
            fb_caption = r[1].strip() if len(r) > 1 else ""
            ig_caption = r[2].strip() if len(r) > 2 else ""
            google_caption = r[3].strip() if len(r) > 3 else ""
            google_action_type = map_action_type(r[4].strip() if len(r) > 4 else "")
            google_action_url = r[5].strip() if len(r) > 5 else ""
            images_raw = r[6].strip() if len(r) > 6 else ""
            
            target_fb = str(r[7]).strip() in ["1", "true", "True", "yes"] if len(r) > 7 and r[7] != "" else True
            target_ig = str(r[8]).strip() in ["1", "true", "True", "yes"] if len(r) > 8 and r[8] != "" else True
            target_story = str(r[9]).strip() in ["1", "true", "True", "yes"] if len(r) > 9 and r[9] != "" else True
            target_google = str(r[10]).strip() in ["1", "true", "True", "yes"] if len(r) > 10 and r[10] != "" else True
            sched_time = r[11].strip() if len(r) > 11 else ""

            # Fallback
            if not ig_caption:
                ig_caption = fb_caption
            if not google_caption:
                google_caption = fb_caption

            images = [img.strip() for img in images_raw.split(",") if img.strip()]
            posts.append({
                "title": title,
                "fb_caption": fb_caption,
                "ig_caption": ig_caption,
                "google_caption": google_caption,
                "google_action_type": google_action_type,
                "google_action_url": google_action_url,
                "images": images,
                "target_fb": target_fb,
                "target_ig": target_ig,
                "target_story": target_story,
                "target_google": target_google,
                "scheduled_time": format_schedule_datetime(sched_time)
            })
    else:
        # Parse Excel (.xlsx, .xls)
        wb = openpyxl.load_workbook(io.BytesIO(file_content), data_only=True)
        ws = wb["Danh Sách Đăng Bài"] if "Danh Sách Đăng Bài" in wb.sheetnames else wb.active
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not any(row):
                continue
            
            # Check column count to support both old and new templates gracefully
            if len(row) >= 11:
                title = str(row[0] or "").strip()
                fb_caption = str(row[1] or "").strip()
                ig_caption = str(row[2] or "").strip()
                google_caption = str(row[3] or "").strip()
                google_action_type = map_action_type(str(row[4] or ""))
                google_action_url = str(row[5] or "").strip()
                images_raw = str(row[6] or "").strip()
                target_fb = str(row[7]).strip() in ["1", "1.0", "true", "True", "yes"] if row[7] is not None and str(row[7]).strip() != "" else True
                target_ig = str(row[8]).strip() in ["1", "1.0", "true", "True", "yes"] if row[8] is not None and str(row[8]).strip() != "" else True
                target_story = str(row[9]).strip() in ["1", "1.0", "true", "True", "yes"] if row[9] is not None and str(row[9]).strip() != "" else True
                target_google = str(row[10]).strip() in ["1", "1.0", "true", "True", "yes"] if row[10] is not None and str(row[10]).strip() != "" else True
                raw_time = row[11] if len(row) > 11 else None
            else:
                # Old template fallback
                title = ""
                fb_caption = str(row[0] or "").strip()
                ig_caption = str(row[1] or "").strip()
                google_caption = fb_caption
                google_action_type = "LEARN_MORE"
                google_action_url = ""
                images_raw = str(row[2] or "").strip()
                target_fb = str(row[3]).strip() in ["1", "1.0", "true", "True", "yes"] if len(row) > 3 and row[3] is not None else True
                target_ig = str(row[4]).strip() in ["1", "1.0", "true", "True", "yes"] if len(row) > 4 and row[4] is not None else True
                target_story = True
                target_google = True
                raw_time = row[5] if len(row) > 5 else None

            # Fallbacks
            if not ig_caption:
                ig_caption = fb_caption
            if not google_caption:
                google_caption = fb_caption

            sched_time = ""
            if isinstance(raw_time, datetime):
                sched_time = raw_time.strftime("%Y-%m-%d %H:%M")
            elif raw_time:
                sched_time = str(raw_time).strip()

            images = [img.strip() for img in images_raw.split(",") if img.strip()]
            posts.append({
                "title": title,
                "fb_caption": fb_caption,
                "ig_caption": ig_caption,
                "google_caption": google_caption,
                "google_action_type": google_action_type,
                "google_action_url": google_action_url,
                "images": images,
                "target_fb": target_fb,
                "target_ig": target_ig,
                "target_story": target_story,
                "target_google": target_google,
                "scheduled_time": format_schedule_datetime(sched_time)
            })
            
    return posts

def format_schedule_datetime(val: str) -> str:
    if not val:
        raise ValueError("Mỗi dòng bulk phải có thời gian lên lịch.")
    return normalize_schedule(val)

def import_bulk_posts(posts: list) -> dict:
    if len(posts) > 200:
        raise ValueError("Mỗi lần chỉ được nhập tối đa 200 bài.")
    for index, post in enumerate(posts, start=1):
        if not any(post.get(target) for target in ("target_fb", "target_ig", "target_google", "target_threads")):
            raise ValueError(f"Dòng {index}: cần chọn ít nhất một nền tảng feed.")
        if post.get("target_ig") and not post.get("images"):
            raise ValueError(f"Dòng {index}: Instagram yêu cầu ít nhất một ảnh.")
        post["scheduled_time"] = normalize_schedule(post.get("scheduled_time", ""))
    created_count = 0
    ids = []
    for p in posts:
        pid = create_post(
            title=p.get("title", ""),
            fb_caption=p.get("fb_caption", ""),
            ig_caption=p.get("ig_caption", ""),
            google_caption=p.get("google_caption", ""),
            threads_caption=p.get("threads_caption", ""),
            threads_topic_tag=p.get("threads_topic_tag", ""),
            images=p.get("images", []),
            target_fb=p.get("target_fb", True),
            target_ig=p.get("target_ig", True),
            target_story=p.get("target_story", True),
            target_google=p.get("target_google", True),
            target_threads=p.get("target_threads", False),
            google_action_type=p.get("google_action_type", "LEARN_MORE"),
            google_action_url=p.get("google_action_url", ""),
            status="scheduled",
            scheduled_time=p.get("scheduled_time")
        )
        ids.append(pid)
        created_count += 1
    return {"success": True, "created_count": created_count, "ids": ids}
